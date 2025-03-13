import sqlite3
import openpyxl
import os
from tkinter import filedialog
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill

def exportar_para_excel():
    try:
        conn = sqlite3.connect(r'app\bd\dados.db')
    except Exception:
        conn = sqlite3.connect(r'G:\Meu Drive\17 - MODELOS\PROGRAMAS\Gerador de Requisições\app\bd\dados.db')
    
    cursor = conn.cursor()
    
    cursor.execute("""
    SELECT nome_usuario, data_criacao, nome_fornecedor, valor, data_criacao, tipo_servico, 
           tipo_pagamento, agencia, os_num, prefixo, contrato
    FROM registros
    WHERE tipo_pagamento != 'FATURAMENTO'
    """)
    
    dados = cursor.fetchall()
    
    conn.close()
    
    # Criando um novo arquivo Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pagamentos"
    
    # Definição das colunas
    colunas = ["RESPONSÁVEL", "DATA", "LANÇAMENTO", "VALOR", "VENCIMENTO", "CONTA CAIXA", 
               "NOTA FISCAL", "TIPO DE PAGAMENTO", "AGENCIA", "OS", "PREFIXO", "CONTRATO", "RECIBO"]
    ws.append(colunas)

    # Definição dos estilos
    font_titulo = Font(name='Calibri Light', size=11, bold=True)
    fill_amarelo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    fill_cinza = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")
    font_padrao = Font(name='Calibri', size=11)

    # Aplicando estilos ao cabeçalho
    for col_num, column_title in enumerate(colunas, 1):
        cell = ws.cell(row=1, column=col_num, value=column_title)
        cell.font = font_titulo
        cell.fill = fill_amarelo if column_title != "RESPONSÁVEL" else fill_cinza

    for linha in dados:
        nome_usuario, data_criacao, nome_fornecedor, valor, vencimento, tipo_servico, \
        tipo_pagamento, agencia, os_num, prefixo, contrato = linha
        
        # Aplicando regras
        conta_caixa = tipo_servico if tipo_servico else "" # verificar regras para todos os tipos de serviços
        agencia = agencia if agencia else "**"
        os_num = os_num if os_num else "**"
        prefixo = prefixo if prefixo else "**"
        
        # Adicionando linha ao Excel
        ws.append([nome_usuario, data_criacao, nome_fornecedor, valor, vencimento,
                   conta_caixa, "", tipo_pagamento, agencia, os_num, prefixo, contrato, ""])
    
    # Ajustando estilos para todas as linhas
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = font_padrao

    # Ajustando largura das colunas
    for col_num, column_title in enumerate(colunas, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 20
    
    ## Salvaondo o arquivo Excel
    caminho_destino = filedialog.askdirectory()

    if not caminho_destino:
        # Usar as duas linhas abaixo após criação da respectiva aba no gerador de Requisições
        #notification_manager = NotificationManager(root)  # passando a instância da janela principal
        #notification_manager.show_notification(f"Nenhum diretório selecionado!", NotifyType.WARNING, bg_color="#404040", text_color="#FFFFFF")
        return

    nome_arquivo_destino = os.path.join(caminho_destino, "pagamentos.xlsx")

    wb.save(nome_arquivo_destino)
    #notification_manager.show_notification(f"Exportação concluída! Arquivo 'pagamentos.xlsx' gerado com sucesso.", NotifyType.SUCCESS, bg_color="#404040", text_color="#FFFFFF")
    print("Exportação concluída! Arquivo 'pagamentos.xlsx' gerado com sucesso.")

# Executar a função de exportação
exportar_para_excel()